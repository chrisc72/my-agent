"""一次性匯入：把凌越 ERP 的原料「一般價」灌進 raw_materials 的成本欄位。

用法：
    python import_erp_prices.py              # dry-run，只印報告不寫入
    python import_erp_prices.py --apply      # 實際寫入資料庫

以「貨品編號」對 raw_materials.ingredient_code 比對。可重複執行，結果相同。

只讀 ERP 檔的「現有使用原料」總表——各字母分頁（C001、C501…）雖然也有
「一般價」欄，但兩邊有 78 筆數字對不起來，且分頁缺單位/淨重/供應商，
故以總表為唯一來源。
"""
from __future__ import annotations

import argparse
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent
DB_PATH = BASE / "data" / "ingredients.db"

ERP_FILE = Path(
    r"D:\OneDrive\03 台灣君代生技有限公司\01 管理部 (行政採購 財務人資 採買)"
    r"\08 ERP 系統\01 凌越ERP\系統編碼\編碼\01 原料 編號 20230425.xlsx"
)
ERP_SHEET = "現有使用原料"
PRICE_SOURCE = f"{ERP_FILE.name} / {ERP_SHEET}"

# ERP 表頭 → 內部欄位。依名稱抓，不寫死欄位字母，ERP 之後改版才不會錯位。
COLUMNS = {
    "code": "貨品編號",
    "name": "貨品名稱",
    "price": "一般價",
    "unit": "基本 單位",
    "net_weight": "淨重",
    "moq": "最低訂購量",
    "supplier": "供應商",
}


def _num(v):
    """轉成數字；空值或非數字回傳 None（不要回 0，0 元和沒填是兩件事）。"""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_erp() -> dict[str, dict]:
    if not ERP_FILE.exists():
        sys.exit(f"找不到 ERP 檔案：{ERP_FILE}")

    # 先複製到暫存再讀：ERP 檔在 OneDrive 上，若正被 Excel 開著或同步中，
    # 直接開會 PermissionError，但複製仍可成功。
    with tempfile.TemporaryDirectory() as tmp:
        local = Path(tmp) / ERP_FILE.name
        shutil.copy2(ERP_FILE, local)
        return _parse_workbook(local)


def _parse_workbook(path: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if ERP_SHEET not in wb.sheetnames:
        sys.exit(f"ERP 檔沒有「{ERP_SHEET}」分頁，實際分頁：{wb.sheetnames}")
    ws = wb[ERP_SHEET]

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {}
    for key, title in COLUMNS.items():
        if title not in header:
            sys.exit(f"ERP 表頭找不到「{title}」欄，實際表頭：{header[:12]}")
        idx[key] = header.index(title)

    out: dict[str, dict] = {}
    for row in rows:
        raw_code = row[idx["code"]]
        if raw_code is None or not str(raw_code).strip():
            continue
        code = str(raw_code).strip()
        out[code] = {
            "name": str(row[idx["name"]] or "").strip(),
            "price": _num(row[idx["price"]]),
            "unit": str(row[idx["unit"]] or "").strip(),
            "net_weight": _num(row[idx["net_weight"]]),
            "moq": _num(row[idx["moq"]]),
            "supplier": str(row[idx["supplier"]] or "").strip(),
        }
    wb.close()
    return out


def main():
    # Windows 主控台預設 cp950，印中文與符號會炸掉
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true",
                    help="實際寫入資料庫（預設只做 dry-run 報告）")
    args = ap.parse_args()

    sys.path.insert(0, str(BASE))
    from database import IngredientDB, cost_per_kg

    erp = read_erp()
    print(f"ERP 來源：{PRICE_SOURCE}")
    print(f"ERP 讀入 {len(erp)} 筆，其中 {sum(1 for v in erp.values() if v['price'] is None)} 筆無價格\n")

    db = IngredientDB(str(DB_PATH))
    materials = db.get_all_materials()

    matched, no_price, unmatched, non_kg = [], [], [], []
    for m in materials:
        code = str(m.get("ingredient_code") or "").strip()
        rec = erp.get(code)
        if not code or rec is None:
            unmatched.append((m["id"], code or "（無編號）", m["product_name"]))
            continue
        if rec["price"] is None:
            no_price.append((code, m["product_name"]))
            continue
        matched.append((code, m["product_name"], rec))
        if cost_per_kg(rec["price"], rec["unit"], rec["net_weight"]) is None:
            non_kg.append((code, rec["unit"], rec["price"]))

    print(f"DB 原料共 {len(materials)} 筆")
    print(f"  ✅ 可寫入價格　　　{len(matched)} 筆")
    print(f"  ⚠️  ERP 有此編號但無價格　{len(no_price)} 筆")
    print(f"  ❌ ERP 找不到此編號　{len(unmatched)} 筆")
    print(f"  ⚠️  單位非 KG 且無淨重、無法換算元/kg　{len(non_kg)} 筆\n")

    if unmatched:
        print("── ERP 找不到的編號（需人工確認，不自動猜測修正）──")
        for mid, code, name in unmatched:
            print(f"   id={mid}  {code}  {name[:50]}")
        print()
    if non_kg:
        print("── 單位非 KG 的項目 ──")
        for code, unit, price in non_kg:
            print(f"   {code}  單位={unit or '（空）'}  一般價={price}")
        print()
    if no_price:
        print(f"── ERP 無價格（保留未設定，不寫 0）：{', '.join(c for c, _ in no_price)}\n")

    if not args.apply:
        print("這是 dry-run，未寫入任何資料。確認以上數字後加 --apply 執行。")
        return

    written = 0
    for code, _name, rec in matched:
        if db.upsert_price_by_code(
            code, rec["price"], rec["unit"], rec["net_weight"],
            rec["moq"], rec["supplier"], PRICE_SOURCE,
        ):
            written += 1
    print(f"✅ 已寫入 {written} 筆原料成本。")


if __name__ == "__main__":
    main()
